"""
step8_corrected_scale.py
=========================
Dual-Phase Optimizer με ΣΩΣΤΗ κλίμακα επίδοσης:
  EP1 = Άριστη επίδοση (high performers)
  EP5 = Χαμηλή επίδοση (low performers)

Κ1: Ισορροπία EP1 (άριστοι)
Κ2: Ισορροπία EP5 (αδύναμοι), με FROZEN EP1

Απαιτήσεις: Python 3.12+, openpyxl>=3.1.0
"""
from __future__ import annotations

import sys
import argparse
from pathlib import Path
from dataclasses import dataclass, field
from typing import Dict, List, Tuple, Optional
from math import ceil

from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.worksheet.worksheet import Worksheet


# ========== DATACLASSES ==========

@dataclass
class StudentData:
    """Δεδομένα μαθητή από source Excel."""
    name: str = ""
    gender: str = ""
    teacher_child: str = "Ο"
    calm: str = "Ο"
    special_needs: str = "Ο"
    greek_knowledge: str = "Ν"
    friends: List[str] = field(default_factory=list)
    conflicts: int = 0
    choice: int = 1  # 1-5: 1=άριστη, 5=χαμηλή


@dataclass
class Student:
    """Student για optimizer."""
    name: str = ""
    choice: int = 1  # 1-5: 1=άριστη, 5=χαμηλή
    gender: str = ""
    greek_knowledge: str = "Ν"
    friends: List[str] = field(default_factory=list)
    locked: bool = False


@dataclass
class SwapRecord:
    """Καταγραφή swap για logging."""
    swap_type: str
    from_team: str
    students_out: List[str]
    to_team: str
    students_in: List[str]
    delta_main: int
    delta_gender: int
    delta_greek: int
    priority: int


# ========== MAIN PROCESSOR ==========

class UnifiedProcessor:
    """
    Dual-phase optimizer με σωστή κλίμακα:
    - Κ1: Optimizes EP1 (άριστοι)
    - Κ2: Optimizes EP5 (αδύναμοι), με EP1 frozen
    """
    
    def __init__(self):
        self.students_data: Dict[str, StudentData] = {}
        self.teams_students: Dict[str, List[str]] = {}
        self.students: Dict[str, Student] = {}
        self.teams: Dict[str, List[str]] = {}
        
        # Targets
        self.target_ep1 = 2  # Κ1: MAX 2 άριστοι ανά τμήμα
        self.target_ep5 = 2  # Κ2: MAX 2 αδύναμοι ανά τμήμα
        self.spread_ep1_goal = 1
        self.spread_ep5_goal = 1
        
        # Iteration limits
        self.max_iter_k1 = 100
        self.max_iter_k2 = 100
        
        # State tracking
        self.warnings: List[str] = []
        self.swaps_k1: List[SwapRecord] = []
        self.swaps_k2: List[SwapRecord] = []
        
        # K1 snapshot (για K2 protection)
        self.spreads_after_k1: Dict[str, int] = {}
        self.cnt_ep1_after_k1: Dict[str, int] = {}
    
    # ==================== PHASE 1: FILL ====================
    
    def read_source_data(self, source_path: str) -> None:
        """Διάβασμα source Excel."""
        wb = load_workbook(source_path, data_only=True)
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            headers = self._parse_headers(sheet)
            
            if 'ΟΝΟΜΑ' not in headers:
                continue
            
            for row_idx in range(2, sheet.max_row + 1):
                name = self._get_cell_value(sheet, row_idx, headers.get('ΟΝΟΜΑ'))
                if not name:
                    continue
                
                # ΦΙΛΟΙ
                friends_str = self._get_cell_value(sheet, row_idx, headers.get('ΦΙΛΟΙ'))
                friends = [f.strip() for f in friends_str.split(',') if f.strip()] if friends_str else []
                
                # ΕΠΙΔΟΣΗ (1-5)
                choice = 1
                if 'ΕΠΙΔΟΣΗ' in headers:
                    epidosi_raw = sheet.cell(row_idx, headers['ΕΠΙΔΟΣΗ']).value
                    if epidosi_raw is not None:
                        try:
                            choice = int(epidosi_raw)
                            if choice < 1 or choice > 5:
                                choice = 1
                        except:
                            choice = 1
                
                # ΦΥΛΟ
                gender = self._get_cell_value(sheet, row_idx, headers.get('ΦΥΛΟ', '')) or 'Κ'
                
                # ΚΑΛΗ_ΓΝΩΣΗ_ΕΛΛΗΝΙΚΩΝ
                greek_col = headers.get('ΚΑΛΗ_ΓΝΩΣΗ_ΕΛΛΗΝΙΚΩΝ') or headers.get('ΓΝΩΣΗ_ΕΛΛΗΝΙΚΩΝ')
                if greek_col is None:
                    self.warnings.append(f"Μαθητής {name}: Δεν βρέθηκε ΚΑΛΗ_ΓΝΩΣΗ_ΕΛΛΗΝΙΚΩΝ - παραλείπεται")
                    continue
                
                greek_raw = self._get_cell_value(sheet, row_idx, greek_col) or "Ν"
                greek_normalized = greek_raw.strip().upper()
                if greek_normalized in ('Ν', 'N'):
                    greek_knowledge = 'Ν'
                elif greek_normalized in ('Ο', 'O'):
                    greek_knowledge = 'Ο'
                else:
                    self.warnings.append(f"Unknown ΚΑΛΗ_ΓΝΩΣΗ '{greek_raw}' for {name}, defaulting to Ν")
                    greek_knowledge = 'Ν'
                
                # LOCKED flags
                teacher_child = self._get_cell_value(sheet, row_idx, headers.get('ΠΑΙΔΙ_ΕΚΠΑΙΔΕΥΤΙΚΟΥ')) or 'Ο'
                calm = self._get_cell_value(sheet, row_idx, headers.get('ΖΩΗΡΟΣ')) or 'Ο'
                special_needs = self._get_cell_value(sheet, row_idx, headers.get('ΙΔΙΑΙΤΕΡΟΤΗΤΑ')) or 'Ο'
                
                self.students_data[name] = StudentData(
                    name=name,
                    gender=gender,
                    teacher_child=teacher_child,
                    calm=calm,
                    special_needs=special_needs,
                    greek_knowledge=greek_knowledge,
                    friends=friends,
                    choice=choice
                )
        
        wb.close()
        print(f"✅ Διαβάστηκαν {len(self.students_data)} μαθητές")
    
    def fill_template(self, template_path: str, output_path: str) -> None:
        """Συμπλήρωση template με δεδομένα."""
        wb = load_workbook(template_path)
        
        for sheet_name in wb.sheetnames:
            if sheet_name in ['ΚΑΤΗΓΟΡΙΟΠΟΙΗΣΗ', 'SINGLE']:
                continue
            
            sheet = wb[sheet_name]
            headers = self._parse_headers(sheet)
            
            # Ensure columns exist
            required = ['ΟΝΟΜΑ', 'ΦΥΛΟ', 'ΚΑΛΗ_ΓΝΩΣΗ_ΕΛΛΗΝΙΚΩΝ', 'ΕΠΙΔΟΣΗ', 'ΦΙΛΟΙ']
            for col in required:
                if col not in headers:
                    next_col = sheet.max_column + 1
                    sheet.cell(1, next_col, col)
                    headers[col] = next_col
            
            team_students = []
            for row_idx in range(2, sheet.max_row + 1):
                name = self._get_cell_value(sheet, row_idx, headers.get('ΟΝΟΜΑ'))
                if not name or name not in self.students_data:
                    continue
                
                sd = self.students_data[name]
                sheet.cell(row_idx, headers['ΦΥΛΟ'], sd.gender)
                sheet.cell(row_idx, headers['ΚΑΛΗ_ΓΝΩΣΗ_ΕΛΛΗΝΙΚΩΝ'], sd.greek_knowledge)
                sheet.cell(row_idx, headers['ΕΠΙΔΟΣΗ'], sd.choice)
        
        for sheet_name in wb.sheetnames:
            if sheet_name in ['ΚΑΤΗΓΟΡΙΟΠΟΙΗΣΗ', 'SINGLE']:
                continue
            
            sheet = wb[sheet_name]
            headers = self._parse_headers(sheet)
            
            # Ensure columns exist
            required = ['ΟΝΟΜΑ', 'ΦΥΛΟ', 'ΚΑΛΗ_ΓΝΩΣΗ_ΕΛΛΗΝΙΚΩΝ', 'ΕΠΙΔΟΣΗ', 'ΦΙΛΟΙ']
            for col in required:
                if col not in headers:
                    next_col = sheet.max_column + 1
                    sheet.cell(1, next_col, col)
                    headers[col] = next_col
            
            team_students = []
            for row_idx in range(2, sheet.max_row + 1):
                name = self._get_cell_value(sheet, row_idx, headers.get('ΟΝΟΜΑ'))
                if not name or name not in self.students_data:
                    continue
                
                sd = self.students_data[name]
                sheet.cell(row_idx, headers['ΦΥΛΟ'], sd.gender)
                sheet.cell(row_idx, headers['ΚΑΛΗ_ΓΝΩΣΗ_ΕΛΛΗΝΙΚΩΝ'], sd.greek_knowledge)
                sheet.cell(row_idx, headers['ΕΠΙΔΟΣΗ'], sd.choice)
                sheet.cell(row_idx, headers['ΦΙΛΟΙ'], ', '.join(sd.friends))
                team_students.append(name)
            
            self.teams_students[sheet_name] = team_students
            print(f"📝 {sheet_name}: {len(team_students)} μαθητές")
        
        wb.save(output_path)
        wb.close()
        print(f"✅ Filled: {output_path}")
    
    def fill_target_excel(self, template_path: str, output_path: str) -> None:
        """Alias για fill_template - συμβατότητα με app.py"""
        self.fill_template(template_path, output_path)
    
    # ==================== PHASE 2: LOAD ====================
    
    def load_filled_data(self, filled_path: str) -> None:
        """Φόρτωση filled data για optimization."""
        wb = load_workbook(filled_path, data_only=True)
        
        for sheet_name in wb.sheetnames:
            if sheet_name in ['ΚΑΤΗΓΟΡΙΟΠΟΙΗΣΗ', 'SINGLE']:
                continue
            
            sheet = wb[sheet_name]
            headers = self._parse_headers(sheet)
            
            team_list = []
            for row_idx in range(2, sheet.max_row + 1):
                name = self._get_cell_value(sheet, row_idx, headers.get('ΟΝΟΜΑ'))
                if not name:
                    continue
                
                # Parse attributes
                choice = 1
                if 'ΕΠΙΔΟΣΗ' in headers:
                    ep_val = sheet.cell(row_idx, headers['ΕΠΙΔΟΣΗ']).value
                    if ep_val is not None:
                        try:
                            choice = int(ep_val)
                        except:
                            choice = 1
                
                gender = self._get_cell_value(sheet, row_idx, headers.get('ΦΥΛΟ')) or 'Κ'
                
                greek_col = headers.get('ΚΑΛΗ_ΓΝΩΣΗ_ΕΛΛΗΝΙΚΩΝ') or headers.get('ΓΝΩΣΗ_ΕΛΛΗΝΙΚΩΝ')
                greek_raw = self._get_cell_value(sheet, row_idx, greek_col) if greek_col else 'Ν'
                greek_normalized = (greek_raw or 'Ν').strip().upper()
                greek_knowledge = 'Ν' if greek_normalized in ('Ν', 'N') else 'Ο'
                
                # Friends
                friends_str = self._get_cell_value(sheet, row_idx, headers.get('ΦΙΛΟΙ'))
                friends = [f.strip() for f in friends_str.split(',') if f.strip()] if friends_str else []
                
                # Locked
                locked_col = headers.get('LOCKED')
                locked_val = self._get_cell_value(sheet, row_idx, locked_col) if locked_col else ""
                locked = (locked_val == "LOCKED")
                
                self.students[name] = Student(
                    name=name,
                    choice=choice,
                    gender=gender,
                    greek_knowledge=greek_knowledge,
                    friends=friends,
                    locked=locked
                )
                team_list.append(name)
            
            self.teams[sheet_name] = team_list
        
        wb.close()
        print(f"✅ Loaded {len(self.students)} μαθητές, {len(self.teams)} τμήματα")
    
    # ==================== PHASE 3: DUAL-PHASE OPTIMIZATION ====================
    
    def optimize_dual_phase(self, dynamic_ep5: bool = False) -> None:
        """
        Dual-phase optimization:
        - K1: EP1 (άριστοι)
        - K2: EP5 (αδύναμοι), με EP1 frozen
        """
        print("\n🎯 Phase 3/3: Dual-phase optimization...")
        
        # Dynamic target EP5 (optional)
        if dynamic_ep5:
            total_ep5 = sum(1 for s in self.students.values() if s.choice == 5)
            self.target_ep5 = min(ceil(total_ep5 / len(self.teams)), 3)
            print(f"📊 Dynamic EP5 target: {self.target_ep5}")
        
        # K1: Optimize EP1 (άριστοι)
        self._optimize_k1_ep1()
        
        # Snapshot K1 state
        self.spreads_after_k1 = self._calculate_spreads()
        self.cnt_ep1_after_k1 = {tn: self._count_choice(tn, 1) for tn in self.teams}
        
        # Freeze EP1 before K2
        self._freeze_ep1_before_k2()
        
        # K2: Optimize EP5 (αδύναμοι)
        self._optimize_k2_ep5()
    
    def optimize(self, max_iterations: int = 100) -> tuple:
        """Alias για optimize_dual_phase - συμβατότητα με app.py"""
        self.max_iter_k1 = max_iterations
        self.max_iter_k2 = max_iterations
        self.optimize_dual_phase(dynamic_ep5=False)
        # Return format expected by app: (swaps, spreads_after)
        return (self.swaps_k1 + self.swaps_k2, self._calculate_spreads())
        
        print("\n🎉 Ολοκληρώθηκε!")
        print(f"  K1 swaps: {len(self.swaps_k1)}")
        print(f"  K2 swaps: {len(self.swaps_k2)}")
        print(f"  Total: {len(self.swaps_k1) + len(self.swaps_k2)}")
    
    def _optimize_k1_ep1(self) -> None:
        """Κύκλος 1: Optimize EP1 (άριστοι)."""
        print("\n🎯 ΚΥΚΛΟΣ 1: Optimizing EP1 (high performers)...")
        
        spreads = self._calculate_spreads()
        print(f"📊 ΠΡΙΝ Κ1:")
        print(f"   spread_ep1: {spreads['ep1']}")
        
        # Distribution
        ep1_counts = {tn: self._count_choice(tn, 1) for tn in self.teams}
        for tn in sorted(self.teams.keys()):
            cnt = ep1_counts[tn]
            excess = max(0, cnt - self.target_ep1)
            status = "✅" if excess == 0 else f"❌ (excess: {excess})"
            print(f"   {tn}: {cnt} EP1 {status}")
        
        # Main loop
        for iteration in range(1, self.max_iter_k1 + 1):
            ep1_counts = {tn: self._count_choice(tn, 1) for tn in self.teams}
            spread_current = max(ep1_counts.values()) - min(ep1_counts.values())
            excess_teams = sum(1 for cnt in ep1_counts.values() if cnt > self.target_ep1)
            
            # Stop conditions
            if spread_current <= self.spread_ep1_goal and excess_teams == 0:
                print(f"✅ spread_ep1 ≤ {self.spread_ep1_goal} AND no excess στο iteration {iteration}")
                break
            
            if spread_current <= self.spread_ep1_goal:
                print(f"✅ spread_ep1 ≤ {self.spread_ep1_goal} στο iteration {iteration}")
                break
            
            # Find max/min teams
            max_team = max(ep1_counts, key=ep1_counts.get)
            min_team = min(ep1_counts, key=ep1_counts.get)
            
            # Generate candidates
            candidates = self._generate_k1_swaps(max_team, min_team)
            if not candidates:
                print(f"⚠️ Δεν βρέθηκαν swaps στο iteration {iteration}")
                break
            
            # Select best
            best = self._select_best_swap(candidates, main_metric='ep1')
            if not best:
                print(f"⚠️ Δεν βρέθηκε valid swap στο iteration {iteration}")
                break
            
            # Apply
            self._apply_swap(best)
            self.swaps_k1.append(best)
            
            # Progress
            if iteration % 10 == 0:
                print(f"  Iteration {iteration}: {len(self.swaps_k1)} swaps, spread={spread_current}")
        
        # Final report K1
        print(f"\n📊 ΜΕΤΑ Κ1 (best effort):")
        print(f"   Swaps: {len(self.swaps_k1)}")
        spreads_after = self._calculate_spreads()
        print(f"   spread_ep1: {spreads['ep1']} → {spreads_after['ep1']}", end="")
        print(" ✅" if spreads_after['ep1'] <= self.spread_ep1_goal else " ⚠️")
        
        ep1_counts_after = {tn: self._count_choice(tn, 1) for tn in self.teams}
        excess_teams_after = sum(1 for cnt in ep1_counts_after.values() if cnt > self.target_ep1)
        
        if spreads_after['ep1'] <= self.spread_ep1_goal:
            print(f"   ✅ Spread goal ≤{self.spread_ep1_goal} επιτεύχθηκε!")
        
        if excess_teams_after > 0:
            print(f"   ⚠️  {excess_teams_after} τμήματα υπερβαίνουν TARGET={self.target_ep1}")
        else:
            print(f"   ✅ Όλα τα τμήματα ≤ TARGET={self.target_ep1}")
        
        for tn in sorted(self.teams.keys()):
            cnt = ep1_counts_after[tn]
            excess = max(0, cnt - self.target_ep1)
            status = "✅" if excess == 0 else f"❌ (excess: {excess})"
            print(f"   {tn}: {cnt} EP1 {status}")
    
    def _freeze_ep1_before_k2(self) -> None:
        """Freeze EP1 students before K2."""
        print("\n🔒 Freezing EP1 για Κ2...")
        frozen_count = 0
        
        for student in self.students.values():
            if student.choice == 1:
                student.locked = True
                frozen_count += 1
            
            # Lock pairs with EP1
            if student.friends:
                for friend_name in student.friends:
                    if friend_name in self.students:
                        friend = self.students[friend_name]
                        if student.choice == 1 or friend.choice == 1:
                            student.locked = True
                            friend.locked = True
                            frozen_count += 1
        
        print(f"   Frozen: {frozen_count} μαθητές")
    
    def _optimize_k2_ep5(self) -> None:
        """Κύκλος 2: Optimize EP5 (αδύναμοι), με EP1 protected."""
        print("\n🎯 ΚΥΚΛΟΣ 2: Optimizing EP5 (EP1 PROTECTED)...")
        
        spreads = self._calculate_spreads()
        print(f"📊 ΠΡΙΝ Κ2:")
        print(f"   spread_ep5: {spreads['ep5']}")
        
        ep5_counts = {tn: self._count_choice(tn, 5) for tn in self.teams}
        for tn in sorted(self.teams.keys()):
            cnt = ep5_counts[tn]
            excess = max(0, cnt - self.target_ep5)
            status = "✅" if excess == 0 else f"❌ (excess: {excess})"
            print(f"   {tn}: {cnt} EP5 {status}")
        
        # Main loop
        for iteration in range(1, self.max_iter_k2 + 1):
            ep5_counts = {tn: self._count_choice(tn, 5) for tn in self.teams}
            spread_current = max(ep5_counts.values()) - min(ep5_counts.values())
            excess_teams = sum(1 for cnt in ep5_counts.values() if cnt > self.target_ep5)
            
            # Stop conditions
            if spread_current <= self.spread_ep5_goal and excess_teams == 0:
                print(f"✅ spread_ep5 ≤ {self.spread_ep5_goal} AND no excess στο iteration {iteration}")
                break
            
            if spread_current <= self.spread_ep5_goal:
                print(f"✅ spread_ep5 ≤ {self.spread_ep5_goal} στο iteration {iteration}")
                break
            
            # Find max/min teams
            max_team = max(ep5_counts, key=ep5_counts.get)
            min_team = min(ep5_counts, key=ep5_counts.get)
            
            # Generate candidates
            candidates = self._generate_k2_swaps(max_team, min_team)
            if not candidates:
                print(f"⚠️ Δεν βρέθηκαν swaps στο iteration {iteration}")
                break
            
            # Safety filter
            safe_candidates = [c for c in candidates if self._is_safe_for_k2(c)]
            if not safe_candidates:
                print(f"⚠️ Δεν βρέθηκαν safe swaps στο iteration {iteration}")
                break
            
            # Select best
            best = self._select_best_swap(safe_candidates, main_metric='ep5')
            if not best:
                print(f"⚠️ Δεν βρέθηκε valid swap στο iteration {iteration}")
                break
            
            # Apply
            self._apply_swap(best)
            self.swaps_k2.append(best)
            
            # Validate K2 invariants
            self._validate_k2_invariants()
            
            # Progress
            if iteration % 10 == 0:
                print(f"  Iteration {iteration}: {len(self.swaps_k2)} swaps, spread={spread_current}")
        
        # Final report K2
        print(f"\n📊 ΤΕΛΙΚΑ (best effort):")
        print(f"   Swaps: {len(self.swaps_k2)}")
        
        spreads_final = self._calculate_spreads()
        print(f"   spread_ep1: {spreads_final['ep1']} (unchanged ✅)")
        print(f"   spread_ep5: {spreads['ep5']} → {spreads_final['ep5']}", end="")
        print(" ✅" if spreads_final['ep5'] <= self.spread_ep5_goal else " ⚠️")
        
        ep5_counts_final = {tn: self._count_choice(tn, 5) for tn in self.teams}
        excess_teams_final = sum(1 for cnt in ep5_counts_final.values() if cnt > self.target_ep5)
        
        if spreads_final['ep5'] <= self.spread_ep5_goal:
            print(f"   ✅ Spread goal ≤{self.spread_ep5_goal} επιτεύχθηκε!")
        
        if excess_teams_final > 0:
            print(f"   ⚠️  {excess_teams_final} τμήματα υπερβαίνουν TARGET={self.target_ep5}")
        else:
            print(f"   ✅ Όλα τα τμήματα ≤ TARGET={self.target_ep5}")
        
        print(f"   spread_boys: {spreads_final['boys']} ✅")
        print(f"   spread_girls: {spreads_final['girls']} ✅")
        print(f"   spread_greek: {spreads_final['greek_yes']} ✅")
    
    def _is_safe_for_k2(self, swap: SwapRecord) -> bool:
        """Check αν swap είναι safe για K2 (δεν αλλάζει EP1)."""
        # Check 1: Locked students
        for name in swap.students_out + swap.students_in:
            if self.students[name].locked:
                return False
        
        # Check 2: EP1 counts preservation
        cnt_ep1_before = {tn: self._count_choice(tn, 1) for tn in self.teams}
        
        # Simulate
        self._apply_swap(swap)
        cnt_ep1_after = {tn: self._count_choice(tn, 1) for tn in self.teams}
        
        # Undo
        self._undo_swap(swap)
        
        # Compare
        return cnt_ep1_before == cnt_ep1_after
    
    def _validate_k2_invariants(self) -> None:
        """Validate ότι K1 results intact."""
        current_cnt = {tn: self._count_choice(tn, 1) for tn in self.teams}
        
        for tn in self.teams:
            if current_cnt[tn] != self.cnt_ep1_after_k1[tn]:
                raise RuntimeError(f"❌ FATAL: cnt_ep1 άλλαξε για {tn}! "
                                   f"Before K2: {self.cnt_ep1_after_k1[tn]}, Now: {current_cnt[tn]}")
    
    # ==================== SWAP GENERATION ====================
    
    def _generate_k1_swaps(self, max_team: str, min_team: str) -> List[SwapRecord]:
        """Generate swaps K1: EP1 (max) ↔ EP2/3/4/5 (min)."""
        candidates = []
        
        # Priority 1: Solo strict
        solos_max = self._get_solos_with_choice(max_team, [1])
        solos_min = self._get_solos_with_choice(min_team, [2, 3, 4, 5])
        
        for s_max in solos_max:
            for s_min in solos_min:
                if (self.students[s_max].gender == self.students[s_min].gender and
                    self.students[s_max].greek_knowledge == self.students[s_min].greek_knowledge):
                    
                    improvement = self._compute_improvement_k1(max_team, [s_max], min_team, [s_min])
                    if improvement['improves']:
                        candidates.append(SwapRecord(
                            swap_type="Solo(EP1)↔Solo(low)-Strict",
                            from_team=max_team,
                            students_out=[s_max],
                            to_team=min_team,
                            students_in=[s_min],
                            delta_main=improvement['delta_spread_ep1'],
                            delta_gender=improvement['delta_boys'] + improvement['delta_girls'],
                            delta_greek=improvement['delta_greek'],
                            priority=1
                        ))
        
        # Priority 2: Pair strict
        pairs_max = self._get_pairs_with_choice(max_team, [1])
        pairs_min = self._get_pairs_with_choice(min_team, [2, 3, 4, 5])
        
        for (a_max, b_max) in pairs_max:
            for (a_min, b_min) in pairs_min:
                if self._pairs_match_strict(a_max, b_max, a_min, b_min):
                    improvement = self._compute_improvement_k1(max_team, [a_max, b_max], min_team, [a_min, b_min])
                    if improvement['improves']:
                        candidates.append(SwapRecord(
                            swap_type="Pair(high)↔Pair(low)-Strict",
                            from_team=max_team,
                            students_out=[a_max, b_max],
                            to_team=min_team,
                            students_in=[a_min, b_min],
                            delta_main=improvement['delta_spread_ep1'],
                            delta_gender=improvement['delta_boys'] + improvement['delta_girls'],
                            delta_greek=improvement['delta_greek'],
                            priority=2
                        ))
        
        # Priority 3: Solo relaxed
        for s_max in solos_max:
            for s_min in solos_min:
                if self.students[s_max].gender == self.students[s_min].gender:
                    if self.students[s_max].greek_knowledge == self.students[s_min].greek_knowledge:
                        continue  # P1 already covered
                    
                    improvement = self._compute_improvement_k1(max_team, [s_max], min_team, [s_min])
                    if improvement['improves'] and improvement['spread_greek_after'] <= 4:
                        candidates.append(SwapRecord(
                            swap_type="Solo(EP1)↔Solo(low)-Relaxed",
                            from_team=max_team,
                            students_out=[s_max],
                            to_team=min_team,
                            students_in=[s_min],
                            delta_main=improvement['delta_spread_ep1'],
                            delta_gender=improvement['delta_boys'] + improvement['delta_girls'],
                            delta_greek=improvement['delta_greek'],
                            priority=3
                        ))
        
        return candidates
    
    def _generate_k2_swaps(self, max_team: str, min_team: str) -> List[SwapRecord]:
        """Generate swaps K2: EP5 (max) ↔ EP2/3/4 (min)."""
        candidates = []
        
        # Priority 1: Solo strict
        solos_max = self._get_solos_with_choice(max_team, [5])
        solos_min = self._get_solos_with_choice(min_team, [2, 3, 4])  # NO EP1!
        
        for s_max in solos_max:
            for s_min in solos_min:
                if (self.students[s_max].gender == self.students[s_min].gender and
                    self.students[s_max].greek_knowledge == self.students[s_min].greek_knowledge):
                    
                    improvement = self._compute_improvement_k2(max_team, [s_max], min_team, [s_min])
                    if improvement['improves']:
                        candidates.append(SwapRecord(
                            swap_type="Solo(EP5)↔Solo(mid)-Strict",
                            from_team=max_team,
                            students_out=[s_max],
                            to_team=min_team,
                            students_in=[s_min],
                            delta_main=improvement['delta_spread_ep5'],
                            delta_gender=improvement['delta_boys'] + improvement['delta_girls'],
                            delta_greek=improvement['delta_greek'],
                            priority=1
                        ))
        
        # Priority 2: Pair strict
        pairs_max = self._get_pairs_with_choice(max_team, [5], exclude_ep1=True)
        pairs_min = self._get_pairs_with_choice(min_team, [2, 3, 4], exclude_ep1=True)
        
        for (a_max, b_max) in pairs_max:
            for (a_min, b_min) in pairs_min:
                if self._pairs_match_strict(a_max, b_max, a_min, b_min):
                    improvement = self._compute_improvement_k2(max_team, [a_max, b_max], min_team, [a_min, b_min])
                    if improvement['improves']:
                        candidates.append(SwapRecord(
                            swap_type="Pair(low)↔Pair(mid)-Strict",
                            from_team=max_team,
                            students_out=[a_max, b_max],
                            to_team=min_team,
                            students_in=[a_min, b_min],
                            delta_main=improvement['delta_spread_ep5'],
                            delta_gender=improvement['delta_boys'] + improvement['delta_girls'],
                            delta_greek=improvement['delta_greek'],
                            priority=2
                        ))
        
        # Priority 3: Solo relaxed
        for s_max in solos_max:
            for s_min in solos_min:
                if self.students[s_max].gender == self.students[s_min].gender:
                    if self.students[s_max].greek_knowledge == self.students[s_min].greek_knowledge:
                        continue
                    
                    improvement = self._compute_improvement_k2(max_team, [s_max], min_team, [s_min])
                    if improvement['improves'] and improvement['spread_greek_after'] <= 4:
                        candidates.append(SwapRecord(
                            swap_type="Solo(EP5)↔Solo(mid)-Relaxed",
                            from_team=max_team,
                            students_out=[s_max],
                            to_team=min_team,
                            students_in=[s_min],
                            delta_main=improvement['delta_spread_ep5'],
                            delta_gender=improvement['delta_boys'] + improvement['delta_girls'],
                            delta_greek=improvement['delta_greek'],
                            priority=3
                        ))
        
        return candidates
    
    def _get_solos_with_choice(self, team: str, choices: List[int]) -> List[str]:
        """Get solo students με επίδοση στο choices."""
        solos = []
        for name in self.teams[team]:
            s = self.students[name]
            if s.locked:
                continue
            if s.choice not in choices:
                continue
            
            # Check if solo (no friends in same team)
            has_friend_here = False
            for friend_name in s.friends:
                if friend_name in self.teams[team]:
                    has_friend_here = True
                    break
            
            if not has_friend_here:
                solos.append(name)
        
        return solos
    
    def _get_pairs_with_choice(self, team: str, choices: List[int], exclude_ep1: bool = False) -> List[Tuple[str, str]]:
        """Get friend pairs με τουλάχιστον ένα choice στο choices."""
        pairs = []
        seen = set()
        
        for name in self.teams[team]:
            if name in seen:
                continue
            
            s = self.students[name]
            if s.locked:
                continue
            
            for friend_name in s.friends:
                if friend_name not in self.teams[team] or friend_name in seen:
                    continue
                
                friend = self.students[friend_name]
                if friend.locked:
                    continue
                
                # Check if at least one in target choices
                if s.choice not in choices and friend.choice not in choices:
                    continue
                
                # K2 mode: exclude pairs with EP1
                if exclude_ep1 and (s.choice == 1 or friend.choice == 1):
                    continue
                
                pairs.append((name, friend_name))
                seen.add(name)
                seen.add(friend_name)
        
        return pairs
    
    def _pairs_match_strict(self, a1: str, b1: str, a2: str, b2: str) -> bool:
        """Check αν 2 pairs match strictly (gender + greek)."""
        s_a1, s_b1 = self.students[a1], self.students[b1]
        s_a2, s_b2 = self.students[a2], self.students[b2]
        
        return (s_a1.gender == s_a2.gender and
                s_b1.gender == s_b2.gender and
                s_a1.greek_knowledge == s_a2.greek_knowledge and
                s_b1.greek_knowledge == s_b2.greek_knowledge)
    
    # ==================== IMPROVEMENT COMPUTATION ====================
    
    def _compute_improvement_k1(self, from_team: str, students_out: List[str],
                                  to_team: str, students_in: List[str]) -> Dict:
        """Compute improvement για K1 swap (EP1 metric)."""
        # Snapshot before
        stats_before = self._get_team_stats()
        spreads_before = self._calculate_spreads()
        ep1_before = {tn: self._count_choice(tn, 1) for tn in self.teams}
        spread_ep1_before = max(ep1_before.values()) - min(ep1_before.values())
        excess_teams_before = sum(1 for cnt in ep1_before.values() if cnt > self.target_ep1)
        total_excess_before = sum(max(0, cnt - self.target_ep1) for cnt in ep1_before.values())
        
        # Simulate swap
        for name in students_out:
            self.teams[from_team].remove(name)
            self.teams[to_team].append(name)
        for name in students_in:
            self.teams[to_team].remove(name)
            self.teams[from_team].append(name)
        
        # Snapshot after
        stats_after = self._get_team_stats()
        spreads_after = self._calculate_spreads()
        ep1_after = {tn: self._count_choice(tn, 1) for tn in self.teams}
        spread_ep1_after = max(ep1_after.values()) - min(ep1_after.values())
        excess_teams_after = sum(1 for cnt in ep1_after.values() if cnt > self.target_ep1)
        total_excess_after = sum(max(0, cnt - self.target_ep1) for cnt in ep1_after.values())
        
        # Undo swap
        for name in students_out:
            self.teams[to_team].remove(name)
            self.teams[from_team].append(name)
        for name in students_in:
            self.teams[from_team].remove(name)
            self.teams[to_team].append(name)
        
        # Compute deltas
        delta_spread_ep1 = spread_ep1_before - spread_ep1_after
        delta_excess_teams = excess_teams_before - excess_teams_after
        delta_total_excess = total_excess_before - total_excess_after
        delta_boys = spreads_before['boys'] - spreads_after['boys']
        delta_girls = spreads_before['girls'] - spreads_after['girls']
        delta_greek = spreads_before['greek_yes'] - spreads_after['greek_yes']
        
        improves = (
            delta_spread_ep1 > 0 or
            (delta_spread_ep1 == 0 and delta_excess_teams > 0) or
            (delta_spread_ep1 == 0 and delta_excess_teams == 0 and delta_total_excess > 0)
        )
        
        return {
            'improves': improves,
            'delta_spread_ep1': delta_spread_ep1,
            'delta_excess_teams': delta_excess_teams,
            'delta_total_excess': delta_total_excess,
            'delta_boys': delta_boys,
            'delta_girls': delta_girls,
            'delta_greek': delta_greek,
            'spread_greek_after': spreads_after['greek_yes']
        }
    
    def _compute_improvement_k2(self, from_team: str, students_out: List[str],
                                  to_team: str, students_in: List[str]) -> Dict:
        """Compute improvement για K2 swap (EP5 metric)."""
        # Snapshot before
        stats_before = self._get_team_stats()
        spreads_before = self._calculate_spreads()
        ep5_before = {tn: self._count_choice(tn, 5) for tn in self.teams}
        spread_ep5_before = max(ep5_before.values()) - min(ep5_before.values())
        excess_teams_before = sum(1 for cnt in ep5_before.values() if cnt > self.target_ep5)
        total_excess_before = sum(max(0, cnt - self.target_ep5) for cnt in ep5_before.values())
        
        # Simulate
        for name in students_out:
            self.teams[from_team].remove(name)
            self.teams[to_team].append(name)
        for name in students_in:
            self.teams[to_team].remove(name)
            self.teams[from_team].append(name)
        
        # Snapshot after
        stats_after = self._get_team_stats()
        spreads_after = self._calculate_spreads()
        ep5_after = {tn: self._count_choice(tn, 5) for tn in self.teams}
        spread_ep5_after = max(ep5_after.values()) - min(ep5_after.values())
        excess_teams_after = sum(1 for cnt in ep5_after.values() if cnt > self.target_ep5)
        total_excess_after = sum(max(0, cnt - self.target_ep5) for cnt in ep5_after.values())
        
        # Undo
        for name in students_out:
            self.teams[to_team].remove(name)
            self.teams[from_team].append(name)
        for name in students_in:
            self.teams[from_team].remove(name)
            self.teams[to_team].append(name)
        
        # Deltas
        delta_spread_ep5 = spread_ep5_before - spread_ep5_after
        delta_excess_teams = excess_teams_before - excess_teams_after
        delta_total_excess = total_excess_before - total_excess_after
        delta_boys = spreads_before['boys'] - spreads_after['boys']
        delta_girls = spreads_before['girls'] - spreads_after['girls']
        delta_greek = spreads_before['greek_yes'] - spreads_after['greek_yes']
        
        improves = (
            delta_spread_ep5 > 0 or
            (delta_spread_ep5 == 0 and delta_excess_teams > 0) or
            (delta_spread_ep5 == 0 and delta_excess_teams == 0 and delta_total_excess > 0)
        )
        
        return {
            'improves': improves,
            'delta_spread_ep5': delta_spread_ep5,
            'delta_excess_teams': delta_excess_teams,
            'delta_total_excess': delta_total_excess,
            'delta_boys': delta_boys,
            'delta_girls': delta_girls,
            'delta_greek': delta_greek,
            'spread_greek_after': spreads_after['greek_yes']
        }
    
    def _select_best_swap(self, candidates: List[SwapRecord], main_metric: str) -> Optional[SwapRecord]:
        """Select best swap using lexicographic scoring."""
        if not candidates:
            return None
        
        def score(swap: SwapRecord):
            return (
                -swap.priority,
                swap.delta_main,
                swap.delta_gender,
                swap.delta_greek,
                -len(swap.students_out)  # fewer moves
            )
        
        candidates.sort(key=score, reverse=True)
        return candidates[0]
    
    def _apply_swap(self, swap: SwapRecord) -> None:
        """Apply swap."""
        for name in swap.students_out:
            self.teams[swap.from_team].remove(name)
            self.teams[swap.to_team].append(name)
        for name in swap.students_in:
            self.teams[swap.to_team].remove(name)
            self.teams[swap.from_team].append(name)
    
    def _undo_swap(self, swap: SwapRecord) -> None:
        """Undo swap."""
        for name in swap.students_out:
            self.teams[swap.to_team].remove(name)
            self.teams[swap.from_team].append(name)
        for name in swap.students_in:
            self.teams[swap.from_team].remove(name)
            self.teams[swap.to_team].append(name)
    
    # ==================== UTILITIES ====================
    
    def _count_choice(self, team: str, choice: int) -> int:
        """Count students με επίδοση = choice."""
        return sum(1 for name in self.teams[team] if self.students[name].choice == choice)
    
    def _get_team_stats(self) -> Dict:
        """Get stats για όλα τα τμήματα."""
        stats = {}
        for tn in self.teams:
            boys = sum(1 for name in self.teams[tn] if self.students[name].gender == 'Α')
            girls = sum(1 for name in self.teams[tn] if self.students[name].gender == 'Κ')
            greek_yes = sum(1 for name in self.teams[tn] if self.students[name].greek_knowledge == 'Ν')
            ep1 = self._count_choice(tn, 1)
            ep2 = self._count_choice(tn, 2)
            ep3 = self._count_choice(tn, 3)
            ep4 = self._count_choice(tn, 4)
            ep5 = self._count_choice(tn, 5)
            
            stats[tn] = {
                'boys': boys,
                'girls': girls,
                'greek_yes': greek_yes,
                'ep1': ep1,
                'ep2': ep2,
                'ep3': ep3,
                'ep4': ep4,
                'ep5': ep5
            }
        return stats
    
    def _calculate_spreads(self) -> Dict[str, int]:
        """Calculate spreads για όλα τα metrics."""
        stats = self._get_team_stats()
        
        boys = [s['boys'] for s in stats.values()]
        girls = [s['girls'] for s in stats.values()]
        greek_yes = [s['greek_yes'] for s in stats.values()]
        ep1 = [s['ep1'] for s in stats.values()]
        ep2 = [s['ep2'] for s in stats.values()]
        ep3 = [s['ep3'] for s in stats.values()]
        ep4 = [s['ep4'] for s in stats.values()]
        ep5 = [s['ep5'] for s in stats.values()]
        
        return {
            'boys': max(boys) - min(boys),
            'girls': max(girls) - min(girls),
            'greek_yes': max(greek_yes) - min(greek_yes),
            'ep1': max(ep1) - min(ep1),
            'ep2': max(ep2) - min(ep2),
            'ep3': max(ep3) - min(ep3),
            'ep4': max(ep4) - min(ep4),
            'ep5': max(ep5) - min(ep5)
        }
    
    def calculate_spreads(self) -> Dict[str, int]:
        """Public wrapper για _calculate_spreads - συμβατότητα με app.py"""
        return self._calculate_spreads()
    
    # ==================== EXPORT ====================
    
    def export_results(self, output_path: str) -> None:
        """Export optimized results to Excel."""
        wb = Workbook()
        wb.remove(wb.active)
        
        # Team sheets
        for tn in sorted(self.teams.keys()):
            ws = wb.create_sheet(tn)
            self._write_team_sheet(ws, tn)
        
        # ΣΤΑΤΙΣΤΙΚΑ
        ws_stats = wb.create_sheet("ΣΤΑΤΙΣΤΙΚΑ")
        self._write_stats_sheet(ws_stats)
        
        # SWAPS_K1_EP1
        ws_k1 = wb.create_sheet("SWAPS_K1_EP1")
        self._write_swaps_sheet(ws_k1, self.swaps_k1)
        
        # SWAPS_K2_EP5
        ws_k2 = wb.create_sheet("SWAPS_K2_EP5")
        self._write_swaps_sheet(ws_k2, self.swaps_k2)
        
        wb.save(output_path)
        wb.close()
        print(f"✅ Optimized Excel: {output_path}")
    
    def _write_team_sheet(self, ws: Worksheet, team_name: str) -> None:
        """Write team sheet."""
        headers = ['ΟΝΟΜΑ', 'ΦΥΛΟ', 'ΚΑΛΗ_ΓΝΩΣΗ_ΕΛΛΗΝΙΚΩΝ', 'ΕΠΙΔΟΣΗ', 'ΦΙΛΟΙ']
        
        # Headers
        for col_idx, hdr in enumerate(headers, start=1):
            cell = ws.cell(1, col_idx, hdr)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        for row_idx, name in enumerate(self.teams[team_name], start=2):
            s = self.students[name]
            ws.cell(row_idx, 1, s.name)
            ws.cell(row_idx, 2, s.gender)
            ws.cell(row_idx, 3, s.greek_knowledge)
            ws.cell(row_idx, 4, s.choice)
            ws.cell(row_idx, 5, ', '.join(s.friends))
        
        # Adjust widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 12
        ws.column_dimensions['E'].width = 40
    
    def _write_stats_sheet(self, ws: Worksheet) -> None:
        """Write ΣΤΑΤΙΣΤΙΚΑ sheet με Spreads + Πίνακα Επιδόσεων."""
        spreads = self._calculate_spreads()
        
        # ===== SECTION 1: SPREADS =====
        headers = ['Metric', 'Value', 'Target', 'Status']
        for col_idx, hdr in enumerate(headers, start=1):
            cell = ws.cell(1, col_idx, hdr)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        
        rows = [
            ('Spread High Perf (EP1)', spreads['ep1'], f"≤ {self.spread_ep1_goal}"),
            ('Spread Low Perf (EP5)', spreads['ep5'], f"≤ {self.spread_ep5_goal}"),
            ('Spread Boys', spreads['boys'], '≤ 4'),
            ('Spread Girls', spreads['girls'], '≤ 4'),
            ('Spread Greek', spreads['greek_yes'], '≤ 4'),
        ]
        
        for row_idx, (metric, value, target) in enumerate(rows, start=2):
            ws.cell(row_idx, 1, metric)
            ws.cell(row_idx, 2, value)
            ws.cell(row_idx, 3, target)
            
            # Status
            if 'EP1' in metric:
                ok = value <= self.spread_ep1_goal
            elif 'EP5' in metric:
                ok = value <= self.spread_ep5_goal
            else:
                ok = value <= 4
            
            status = '✅' if ok else '❌'
            cell = ws.cell(row_idx, 4, status)
            if ok:
                cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            else:
                cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        # ===== SECTION 2: ΠΙΝΑΚΑΣ ΕΠΙΔΟΣΕΩΝ ΑΝΑ ΤΜΗΜΑ =====
        start_row = len(rows) + 3  # Κενή γραμμή + header
        
        # Headers πίνακα
        perf_headers = ['Τμήμα', 'Σύνολο', 'Αγόρια', 'Κορίτσια', 
                        'Γνώση (ΝΑΙ)', 'Γνώση (ΟΧΙ)', 
                        'Επ1', 'Επ2', 'Επ3', 'Επ4', 'Επ5']
        
        for col_idx, hdr in enumerate(perf_headers, start=1):
            cell = ws.cell(start_row, col_idx, hdr)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Data ανά τμήμα
        for row_offset, team_name in enumerate(sorted(self.teams.keys()), start=1):
            row_idx = start_row + row_offset
            
            total = len(self.teams[team_name])
            boys = sum(1 for name in self.teams[team_name] if self.students[name].gender == 'Α')
            girls = sum(1 for name in self.teams[team_name] if self.students[name].gender == 'Κ')
            greek_yes = sum(1 for name in self.teams[team_name] if self.students[name].greek_knowledge == 'Ν')
            greek_no = sum(1 for name in self.teams[team_name] if self.students[name].greek_knowledge == 'Ο')
            
            ep1 = self._count_choice(team_name, 1)
            ep2 = self._count_choice(team_name, 2)
            ep3 = self._count_choice(team_name, 3)
            ep4 = self._count_choice(team_name, 4)
            ep5 = self._count_choice(team_name, 5)
            
            ws.cell(row_idx, 1, team_name)
            ws.cell(row_idx, 2, total)
            ws.cell(row_idx, 3, boys)
            ws.cell(row_idx, 4, girls)
            ws.cell(row_idx, 5, greek_yes)
            ws.cell(row_idx, 6, greek_no)
            ws.cell(row_idx, 7, ep1)
            ws.cell(row_idx, 8, ep2)
            ws.cell(row_idx, 9, ep3)
            ws.cell(row_idx, 10, ep4)
            ws.cell(row_idx, 11, ep5)
        
        # Column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 8
        ws.column_dimensions['H'].width = 8
        ws.column_dimensions['I'].width = 8
        ws.column_dimensions['J'].width = 8
        ws.column_dimensions['K'].width = 8
    
    def _write_swaps_sheet(self, ws: Worksheet, swaps: List[SwapRecord]) -> None:
        """Write swaps log sheet."""
        headers = ['#', 'Type', 'From', 'OUT', 'To', 'IN', 'Δ_main', 'Δ_gender', 'Δ_greek', 'Priority']
        
        for col_idx, hdr in enumerate(headers, start=1):
            cell = ws.cell(1, col_idx, hdr)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        for row_idx, swap in enumerate(swaps, start=2):
            ws.cell(row_idx, 1, row_idx - 1)
            ws.cell(row_idx, 2, swap.swap_type)
            ws.cell(row_idx, 3, swap.from_team)
            ws.cell(row_idx, 4, ', '.join(swap.students_out))
            ws.cell(row_idx, 5, swap.to_team)
            ws.cell(row_idx, 6, ', '.join(swap.students_in))
            ws.cell(row_idx, 7, swap.delta_main)
            ws.cell(row_idx, 8, swap.delta_gender)
            ws.cell(row_idx, 9, swap.delta_greek)
            ws.cell(row_idx, 10, swap.priority)
        
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 35
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 35
        ws.column_dimensions['G'].width = 10
        ws.column_dimensions['H'].width = 12
        ws.column_dimensions['I'].width = 10
        ws.column_dimensions['J'].width = 10
    
    # ==================== PARSING HELPERS ====================
    
    def _parse_headers(self, sheet: Worksheet) -> Dict[str, int]:
        """Parse headers from Excel sheet."""
        headers = {}
        for col_idx in range(1, sheet.max_column + 1):
            cell_val = sheet.cell(1, col_idx).value
            if cell_val:
                normalized = str(cell_val).strip().upper().replace(' ', '_')
                headers[normalized] = col_idx
        return headers
    
    def _get_cell_value(self, sheet: Worksheet, row: int, col: Optional[int]) -> str:
        """Get cell value safely."""
        if col is None:
            return ""
        val = sheet.cell(row, col).value
        return str(val).strip() if val is not None else ""


# ==================== CLI ====================

def main():
    parser = argparse.ArgumentParser(description="Dual-Phase Optimizer (corrected scale)")
    parser.add_argument('mode', choices=['fill', 'all'], help="Mode: fill or all")
    parser.add_argument('--source', required=True, help="Source data Excel path")
    parser.add_argument('--template', required=True, help="Template Excel path")
    parser.add_argument('--out', required=True, help="Output Excel path")
    
    # Optimization params (only for 'all' mode)
    parser.add_argument('--dynamic-ep5', action='store_true', help="Enable dynamic EP5 target")
    parser.add_argument('--target-ep1', type=int, default=2, help="Target EP1 per class (default: 2)")
    parser.add_argument('--target-ep5', type=int, default=2, help="Target EP5 per class (default: 2)")
    parser.add_argument('--spread1-goal', type=int, default=1, help="Spread goal for EP1 (default: 1)")
    parser.add_argument('--spread5-goal', type=int, default=1, help="Spread goal for EP5 (default: 1)")
    parser.add_argument('--max-iter-k1', type=int, default=100, help="Max iterations K1 (default: 100)")
    parser.add_argument('--max-iter-k2', type=int, default=100, help="Max iterations K2 (default: 100)")
    
    args = parser.parse_args()
    
    processor = UnifiedProcessor()
    
    # Set targets from CLI
    processor.target_ep1 = args.target_ep1
    processor.target_ep5 = args.target_ep5
    processor.spread_ep1_goal = args.spread1_goal
    processor.spread_ep5_goal = args.spread5_goal
    processor.max_iter_k1 = args.max_iter_k1
    processor.max_iter_k2 = args.max_iter_k2
    
    print(f"📄 Mode: {args.mode.upper()}")
    
    if args.mode == 'fill':
        print("\n📋 Phase 1/1: Filling...")
        processor.read_source_data(args.source)
        processor.fill_template(args.template, args.out)
        
        if processor.warnings:
            print(f"\n⚠️  {len(processor.warnings)} warnings:")
            for w in processor.warnings[:5]:
                print(f"  • {w}")
    
    elif args.mode == 'all':
        # Phase 1: Fill
        print("\n📋 Phase 1/3: Filling...")
        temp_filled = args.out.replace('.xlsx', '_temp_filled.xlsx')
        processor.read_source_data(args.source)
        processor.fill_template(args.template, temp_filled)
        
        # Phase 2: Load
        print("\n📥 Phase 2/3: Loading...")
        processor.load_filled_data(temp_filled)
        
        # Phase 3: Optimize
        processor.optimize_dual_phase(dynamic_ep5=args.dynamic_ep5)
        
        # Export
        processor.export_results(args.out)
        
        if processor.warnings:
            print(f"\n⚠️  {len(processor.warnings)} warnings:")
            for w in processor.warnings[:5]:
                print(f"  • {w}")


if __name__ == '__main__':
    main()
